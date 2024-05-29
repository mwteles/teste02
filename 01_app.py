import streamlit as st
import pandas as pd
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, PatternFill, NamedStyle, Protection, Alignment

st.set_page_config(layout='wide')

def copy_cell_style(source_cell, target_cell):
    target_cell.font = Font(name=source_cell.font.name,
                            size=source_cell.font.size,
                            bold=source_cell.font.bold,
                            italic=source_cell.font.italic,
                            vertAlign=source_cell.font.vertAlign,
                            underline=source_cell.font.underline,
                            strike=source_cell.font.strike,
                            color=source_cell.font.color)
    target_cell.border = Border(left=source_cell.border.left,
                                right=source_cell.border.right,
                                top=source_cell.border.top,
                                bottom=source_cell.border.bottom,
                                diagonal=source_cell.border.diagonal,
                                diagonal_direction=source_cell.border.diagonal_direction,
                                outline=source_cell.border.outline,
                                vertical=source_cell.border.vertical,
                                horizontal=source_cell.border.horizontal)
    target_cell.fill = PatternFill(fill_type=source_cell.fill.fill_type,
                                   start_color=source_cell.fill.start_color,
                                   end_color=source_cell.fill.end_color)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = Protection(locked=source_cell.protection.locked,
                                        hidden=source_cell.protection.hidden)
    target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                      vertical=source_cell.alignment.vertical,
                                      text_rotation=source_cell.alignment.text_rotation,
                                      wrap_text=source_cell.alignment.wrap_text,
                                      shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                      indent=source_cell.alignment.indent)

consolidar = st.sidebar.button('Atualizar')

if consolidar:
    try:
        # Caminhos dos arquivos Excel
        arquivo_status = 'STATUS.xlsx'
        arquivo_transporte = 'TRANSPORTE.xlsx'
        novo_arquivo = 'ACOMPANHAMENTO.xlsx'

        # Ler as abas específicas dos arquivos Excel
        dados_dashboard = pd.read_excel(arquivo_status, sheet_name='DASHBOARD')
        dados_resumodesloc = pd.read_excel(arquivo_transporte, sheet_name='RESUMODESLOC')

        # Salvar os dados em um novo arquivo Excel mantendo a formatação
        with pd.ExcelWriter(novo_arquivo, engine='openpyxl') as writer:
            # Escrever a aba DASHBOARD
            dados_dashboard.to_excel(writer, sheet_name='DASHBOARD', index=False)
            # Escrever a aba RESUMODESLOC
            dados_resumodesloc.to_excel(writer, sheet_name='RESUMODESLOC', index=False)

        # Carregar o novo arquivo para aplicar a formatação das abas de origem
        novo_wb = load_workbook(novo_arquivo)

        # Carregar os arquivos de origem para copiar a formatação
        wb_status = load_workbook(arquivo_status)
        wb_transporte = load_workbook(arquivo_transporte)

        # Copiar a formatação da aba DASHBOARD
        orig_dashboard = wb_status['DASHBOARD']
        new_dashboard = novo_wb['DASHBOARD']
        for row in orig_dashboard.iter_rows():
            for cell in row:
                new_cell = new_dashboard[cell.coordinate]
                copy_cell_style(cell, new_cell)

        # Copiar a formatação da aba RESUMODESLOC
        orig_resumodesloc = wb_transporte['RESUMODESLOC']
        new_resumodesloc = novo_wb['RESUMODESLOC']
        for row in orig_resumodesloc.iter_rows():
            for cell in row:
                new_cell = new_resumodesloc[cell.coordinate]
                copy_cell_style(cell, new_cell)

        # Salvar o novo arquivo com a formatação copiada
        novo_wb.save(novo_arquivo)
       # st.success(f'O arquivo {novo_arquivo} foi criado com sucesso e a formatação foi mantida.')

    except Exception as e:
        st.error(f"Ocorreu um erro: {e}")

st.subheader('Contrato de manutenção predial (17/2023)',divider='blue')
# Upload de arquivo para visualização
uploaded_file = st.file_uploader("APÓS ATUALIZAR, CARREGUE O ARQUIVO: ''ACOMPANHAMENTO.xlsx''", type='.xlsx')

if uploaded_file is not None:
    aba = st.sidebar.selectbox('Selecione análise:',['VISÃO GERAL','DEMANDAS','TRANSPORTE'])
    status = pd.read_excel(uploaded_file, sheet_name='DASHBOARD')
    status['ANO']=status['ANO'].astype(str)
    ano = st.sidebar.radio('Ano:',status['ANO'].unique())
    status = status.loc[status['ANO']==ano]

    encerrados = status.loc[status['STATUS']=='ENCERRADO']
    abertos = status.loc[status['STATUS']!='ENCERRADO']
   
    if aba=='VISÃO GERAL':
       finalizar = {'ENTREGA PARCIAL':'A ENCERRAR','PENDENTE':'A ENCERRAR','INICIADO':'A ENCERRAR','AGENDADO':'A ENCERRAR'}
       abertos['STATUS']=abertos['STATUS'].map(finalizar)
       comparativo = pd.concat([encerrados[['STATUS','MÊS','ANO','CONTROLE','OS']],abertos[['STATUS','MÊS','ANO','CONTROLE','OS']]],ignore_index=True)
       comparativo
    
    if aba=='DEMANDAS':
        status